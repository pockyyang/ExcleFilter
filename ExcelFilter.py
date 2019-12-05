# -*- coding: utf-8 -*-
import openpyxl
x = 1
y = 1
r = 1
c = 1
#file1.xlsx is used to  store the data
#Extract data from file2.xlsx
wb1 = openpyxl.load_workbook('file1.xlsx')
wb2 = openpyxl.load_workbook('myfile1.xlsx')
datasheet = wb1["Sheet1"]
for sheet in wb2:
    for col in sheet.iter_cols(min_row=1, max_col=60, max_row=60):
        for cell in col:
            v = cell.value
            r = cell.row
            c = cell.column
            r_next = r
            c_next = c + 1
            v_next = sheet.cell(row=r_next, column=c_next).value
            if ((v=='C')&(v_next!='')&(v_next!='已解决')):
              x = 1
              for x in  range(1, 40, 1):
                datasheet.cell(row=y, column=x).value = sheet.cell(row=r, column=x).value
              y = y+1

wb1.save('file1.xlsx')
wb2.save('myfile1.xlsx')
